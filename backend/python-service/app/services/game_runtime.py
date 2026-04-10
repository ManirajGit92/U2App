import asyncio
from copy import deepcopy
from datetime import UTC, datetime
from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.schemas.game import (
    GameControlRequest,
    GameStateResponse,
    PhonebookUpsert,
    PublicQuestion,
    QuestionConfig,
    SmsLogEntry,
    WinnerSchema,
)


class GameRuntimeService:
    def __init__(self) -> None:
        self._lock = asyncio.Lock()
        self.broadcast_callback = None
        self._timer_task: asyncio.Task | None = None
        self.questions: list[QuestionConfig] = []
        self.phonebook: dict[str, str] = {}
        self.status = "idle"
        self.current_index = -1
        self.current_question: QuestionConfig | None = None
        self.timer_remaining = 0
        self.break_remaining = 0
        self.scores: dict[str, int] = {}
        self.sms_log: list[SmsLogEntry] = []
        self.winner: WinnerSchema | None = None

    async def reset_runtime(self) -> None:
        async with self._lock:
            await self._cancel_timer()
            self.questions = []
            self.phonebook = {}
            self.status = "idle"
            self.current_index = -1
            self.current_question = None
            self.timer_remaining = 0
            self.break_remaining = 0
            self.scores = {}
            self.sms_log = []
            self.winner = None

    async def control(self, payload: GameControlRequest) -> dict[str, str | bool]:
        async with self._lock:
            if payload.action == "start":
                if not self.questions:
                    raise ValueError("No questions loaded")
                self.scores = {}
                self.sms_log = []
                self.current_index = -1
                await self._next_question()
            elif payload.action == "next":
                await self._cancel_timer()
                if payload.skip_break or not self.current_question:
                    await self._next_question()
                else:
                    await self._start_break_timer(self.current_question.break_after_win)
            elif payload.action == "pause":
                await self._cancel_timer()
                self.status = "paused"
                await self._broadcast("game_state", self.get_public_state().model_dump(mode="json"))
            elif payload.action == "resume":
                if self.current_question and self.timer_remaining > 0:
                    self.status = "playing"
                    self._timer_task = asyncio.create_task(self._question_timer(self.timer_remaining))
                    await self._broadcast("game_state", self.get_public_state().model_dump(mode="json"))
            elif payload.action == "reset":
                await self._cancel_timer()
                self.status = "idle"
                self.current_index = -1
                self.current_question = None
                self.scores = {}
                self.sms_log = []
                self.winner = None
                self.timer_remaining = 0
                self.break_remaining = 0
                await self._broadcast("game_reset", {})
                await self._broadcast("game_state", self.get_public_state().model_dump(mode="json"))
            return {"ok": True, "status": self.status}

    async def load_excel_config(self, file_bytes: bytes) -> tuple[int, int]:
        workbook = load_workbook(BytesIO(file_bytes))
        questions: list[QuestionConfig] = []
        phonebook: dict[str, str] = {}

        if "Questions" in workbook.sheetnames:
            rows = list(workbook["Questions"].iter_rows(values_only=True))
            if rows:
                headers = [str(cell) if cell is not None else "" for cell in rows[0]]
                for row in rows[1:]:
                    mapped = dict(zip(headers, row))
                    answer = str(mapped.get("Answer") or "").strip()
                    if not answer:
                        continue
                    questions.append(
                        QuestionConfig(
                            no=int(mapped.get("No") or len(questions) + 1),
                            content_url=str(mapped.get("ContentURL") or ""),
                            content_type=str(mapped.get("ContentType") or "image").lower(),
                            answer=answer,
                            timer=int(mapped.get("Timer(s)") or 60),
                            points=int(mapped.get("Points") or 10),
                            breakAfterWin=int(mapped.get("BreakAfterWin(s)") or 5),
                            hint=str(mapped.get("Hint") or ""),
                        )
                    )

        if "Phonebook" in workbook.sheetnames:
            rows = list(workbook["Phonebook"].iter_rows(values_only=True))
            if rows:
                headers = [str(cell) if cell is not None else "" for cell in rows[0]]
                for row in rows[1:]:
                    mapped = dict(zip(headers, row))
                    phone = str(mapped.get("Phone") or "").strip()
                    if phone:
                        phonebook[phone] = str(mapped.get("PlayerName") or phone).strip()

        async with self._lock:
            self.questions = questions
            self.phonebook = phonebook
            if self.status == "idle":
                self.current_index = -1
                self.current_question = None

        await self._broadcast("config_loaded", {"questionCount": len(questions), "phonebookCount": len(phonebook)})
        return len(questions), len(phonebook)

    async def handle_sms_answer(self, phone: str, message: str) -> None:
        async with self._lock:
            normalized_phone = self._normalize_phone(phone)
            name = self.phonebook.get(normalized_phone, self.phonebook.get(phone, phone))
            entry = SmsLogEntry(phone=phone, name=name, message=message, time=datetime.now(UTC), correct=False)
            if self.status != "playing" or not self.current_question:
                self.sms_log.append(entry)
                await self._broadcast("sms_received", entry.model_dump(mode="json"))
                return

            is_correct = message.strip().lower() == self.current_question.answer.strip().lower()
            entry.correct = is_correct
            self.sms_log.append(entry)
            await self._broadcast("sms_received", entry.model_dump(mode="json"))

            if is_correct:
                points = self.current_question.points
                self.scores[name] = self.scores.get(name, 0) + points
                self.winner = WinnerSchema(name=name, phone=phone, points=points, message=message)
                self.status = "paused"
                await self._cancel_timer()
                await self._broadcast("winner", {"winner": self.winner.model_dump(mode="json"), "scores": self.scores})
                await self._broadcast("game_state", self.get_public_state().model_dump(mode="json"))
                self._timer_task = asyncio.create_task(self._deferred_break(self.current_question.break_after_win, 4))

    async def upsert_phonebook(self, payload: PhonebookUpsert) -> None:
        async with self._lock:
            self.phonebook[payload.phone.strip()] = payload.name.strip()

    def get_phonebook(self) -> dict[str, str]:
        return deepcopy(self.phonebook)

    def get_public_state(self) -> GameStateResponse:
        question = None
        if self.current_question:
            show_answer = self.status in {"paused", "break", "finished"}
            question = PublicQuestion(
                no=self.current_question.no,
                content_url=self.current_question.content_url,
                content_type=self.current_question.content_type,
                answer=self.current_question.answer if show_answer else None,
                timer=self.current_question.timer,
                points=self.current_question.points,
                break_after_win=self.current_question.break_after_win,
                hint=self.current_question.hint,
            )
        return GameStateResponse(
            status=self.status,
            current_index=self.current_index,
            total_questions=len(self.questions),
            current_question=question,
            timer_remaining=self.timer_remaining,
            break_remaining=self.break_remaining,
            scores=deepcopy(self.scores),
            sms_log=self.sms_log[-20:],
            winner=self.winner,
        )

    def build_template_workbook(self) -> bytes:
        workbook = Workbook()
        questions_sheet = workbook.active
        questions_sheet.title = "Questions"
        questions_sheet.append(["No", "ContentURL", "ContentType", "Answer", "Timer(s)", "Points", "BreakAfterWin(s)", "Hint"])
        questions_sheet.append([1, "https://example.com/dog.jpg", "image", "Dog", 60, 10, 5, "Man's best friend"])
        questions_sheet.append([2, "https://example.com/eiffel.jpg", "image", "Eiffel Tower", 60, 15, 5, "Famous landmark in France"])

        phonebook_sheet = workbook.create_sheet("Phonebook")
        phonebook_sheet.append(["Phone", "PlayerName"])
        phonebook_sheet.append(["+911234567890", "Alice"])
        phonebook_sheet.append(["+910987654321", "Bob"])

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def build_leaderboard_workbook(self) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Leaderboard"
        sheet.append(["Rank", "Player", "Score"])
        for index, (name, score) in enumerate(sorted(self.scores.items(), key=lambda item: item[1], reverse=True), start=1):
            sheet.append([index, name, score])
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    async def _next_question(self) -> None:
        self.current_index += 1
        if self.current_index >= len(self.questions):
            self.status = "finished"
            self.current_question = None
            await self._cancel_timer()
            await self._broadcast("game_over", {"scores": self.scores})
            await self._broadcast("game_state", self.get_public_state().model_dump(mode="json"))
            return
        self.current_question = self.questions[self.current_index]
        self.winner = None
        self.status = "playing"
        await self._broadcast("question", self.current_question.model_dump(by_alias=True))
        await self._broadcast("game_state", self.get_public_state().model_dump(mode="json"))
        self._timer_task = asyncio.create_task(self._question_timer(self.current_question.timer))

    async def _question_timer(self, seconds: int) -> None:
        self.timer_remaining = seconds
        await self._broadcast("timer", {"remaining": self.timer_remaining, "total": seconds})
        while self.timer_remaining > 0:
            await asyncio.sleep(1)
            self.timer_remaining -= 1
            await self._broadcast("timer", {"remaining": self.timer_remaining, "total": seconds})
        self.status = "paused"
        await self._broadcast(
            "time_up",
            {"question": self.current_question.model_dump(by_alias=True) if self.current_question else None, "answer": self.current_question.answer if self.current_question else None},
        )
        await self._broadcast("game_state", self.get_public_state().model_dump(mode="json"))
        if self.current_question:
            self._timer_task = asyncio.create_task(self._deferred_break(5, 3))

    async def _start_break_timer(self, seconds: int) -> None:
        await self._cancel_timer()
        self.break_remaining = seconds
        self.status = "break"
        await self._broadcast("break_start", {"remaining": seconds})
        await self._broadcast("game_state", self.get_public_state().model_dump(mode="json"))
        self._timer_task = asyncio.create_task(self._break_timer(seconds))

    async def _break_timer(self, seconds: int) -> None:
        self.break_remaining = seconds
        while self.break_remaining > 0:
            await asyncio.sleep(1)
            self.break_remaining -= 1
            await self._broadcast("break_timer", {"remaining": self.break_remaining})
        await self._next_question()

    async def _deferred_break(self, break_seconds: int, wait_seconds: int) -> None:
        await asyncio.sleep(wait_seconds)
        if self.status == "paused":
            await self._start_break_timer(break_seconds)

    async def _cancel_timer(self) -> None:
        if self._timer_task and not self._timer_task.done():
            self._timer_task.cancel()
            try:
                await self._timer_task
            except asyncio.CancelledError:
                pass
        self._timer_task = None

    async def _broadcast(self, event: str, payload: dict) -> None:
        if self.broadcast_callback:
            await self.broadcast_callback(event, payload)

    @staticmethod
    def _normalize_phone(phone: str) -> str:
        return "".join(character for character in phone if character not in {" ", "-", "(", ")"})


game_runtime = GameRuntimeService()
